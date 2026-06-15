"""Generate supabase/seed.sql from the real Excel source files."""
import openpyxl, json, re, sys
from typing import Optional

QUESTIONS_FILE = 'C:/Users/Berk/Desktop/Question Master_Berlin.xlsx'
CARE_HOMES_FILE = 'C:/Users/Berk/Desktop/seed-care homes.xlsx'
PLZ_FILE       = 'C:/Users/Berk/Desktop/plz_de.xlsx'

# ── UUID helpers ──────────────────────────────────────────────────────────────
def suuid(n):   return f'10000000-0000-0000-0000-{n:012x}'
def cuuid(n):   return f'20000000-0000-0000-0000-{n:012x}'
def quuid(n):   return f'30000000-0000-0000-0000-{n:012x}'
def catuuid(n): return f'40000000-0000-0000-0000-{n:012x}'
def guuid(n):   return f'50000000-0000-0000-0000-{n:012x}'
def qquid(n):   return f'60000000-0000-0000-0000-{n:012x}'
def ouuid(q,o): return f'70000000-0000-0000-{q:04x}-{o:012x}'
def plzuuid(i): return f'a0000000-0000-0000-0000-{i:012x}'

# ── SQL helpers ───────────────────────────────────────────────────────────────
def sq(s):   return 'NULL' if s is None else "'" + str(s).replace("'", "''") + "'"
def jb(d):   return 'NULL' if d is None else "'" + json.dumps(d, ensure_ascii=False).replace("'","''") + "'::jsonb"

# ── Type mapping ──────────────────────────────────────────────────────────────
TYPE_MAP = {
    'short text':                   'short_text',
    'DD.MM.YYYY':                   'date',
    'DD.MM.YYYY ':                  'date',
    'YYYY':                         'short_text',
    'selection':                    'single_select',
    'multiple_selection':           'multi_select',
    'Euro':                         'amount',
    'number':                       'number',
    '5-digit-number':               'short_text',
    'lookup based on last address PLZ': None,   # auto – skip
}

# ── Key normalisation (fix typos in source) ───────────────────────────────────
KEY_FIX = {
    'maritial_status':                  'marital_status',
    'maritial_status_since':            'marital_status_since',
    'disablity_card':                   'disability_card',
    'disablity card_expiry':            'disability_card_expiry',
    'disablity card_application':       'disability_card_application',
    'child_maritial_status':            'child_marital_status',
    'spouse_disablity_card':            'spouse_disability_card',
    'spouse_disablity card_expiry':     'spouse_disability_card_expiry',
    'spouse_disablity_card_application':'spouse_disability_card_application',
    'spouse_maritial_status':           'spouse_marital_status',
}

# Keys referenced in dependency strings also need normalisation
DEP_KEY_FIX = {
    **KEY_FIX,
    'wohngeld_amount':                  'wohngeld_yes_no',
    'rent_contract_termination':        'rent_contract_termination_yes_no',
    'disability_card':                  'disability_card',
    'spouse_disability_card':           'spouse_disability_card',
}

def parse_dep(dep: Optional[str]) -> Optional[dict]:
    if not dep or not dep.strip():
        return None
    dep = dep.strip()
    # "if X = Y"
    m = re.match(r'if (\S+) = (.+)', dep)
    if m:
        k = DEP_KEY_FIX.get(m.group(1), m.group(1))
        return {"question_key": k, "value": m.group(2).strip()}
    # "if X other than Y"
    m = re.match(r'if (\S+) other than (.+)', dep)
    if m:
        k = DEP_KEY_FIX.get(m.group(1), m.group(1))
        return {"question_key": k, "not_value": m.group(2).strip()}
    # "if X not empty"
    m = re.match(r'if (\S+) not empty', dep)
    if m:
        k = DEP_KEY_FIX.get(m.group(1), m.group(1))
        return {"question_key": k, "not_empty": True}
    return None

# ── Categories ────────────────────────────────────────────────────────────────
SECTIONS = {
    'Personal':'personal', 'Home':'home', 'Children':'children',
    'Income':'income', 'Expenditure':'expenditure', 'Wealth':'wealth',
    'Additional':'additional', 'Spouse':'spouse',
}
CAT_SORT  = {'personal':0,'home':1,'children':2,'income':3,
             'expenditure':4,'wealth':5,'additional':6,'spouse':7}
CAT_LABEL = {
    'personal':'Persönliche Angaben','home':'Wohnverhältnisse',
    'children':'Kinder','income':'Einkünfte','expenditure':'Ausgaben',
    'wealth':'Vermögen','additional':'Weitere Angaben',
    'spouse':'Ehepartner / Lebenspartner',
}
CAT_ID = {s: catuuid(i+1) for i,s in enumerate(sorted(CAT_SORT, key=CAT_SORT.get))}

# ── Repeatable question groups ────────────────────────────────────────────────
GROUPS = {
    'children': {
        'gid':guuid(1),'cat':'children','sort':0,'rep':True,
        'label':'Kinder','key':'children',
        'keys':{'child_first_name','child_last_name','child_birth_name',
                'child_birth_date','child_marital_status','child_family_tie',
                'child_profession','child_address'},
    },
    'pension': {
        'gid':guuid(2),'cat':'income','sort':0,'rep':True,
        'label':'Rente / Pension','key':'pension',
        'keys':{'pension_type','pension_amount','pension_id','pension_issuer'},
    },
    'other_income': {
        'gid':guuid(3),'cat':'income','sort':1,'rep':True,
        'label':'Sonstige Einkünfte','key':'other_income',
        'keys':{'other_income_type','other_income_amount'},
    },
    'bank_additional': {
        'gid':guuid(4),'cat':'wealth','sort':0,'rep':True,
        'label':'Weitere Bankkonten','key':'bank_additional',
        'keys':{'bank_additional_name','bank_additional_iban','bank_additional_amount'},
    },
    'spouse_pension': {
        'gid':guuid(5),'cat':'spouse','sort':0,'rep':True,
        'label':'Rente / Pension des Ehepartners','key':'spouse_pension',
        'keys':{'spouse_pension_type','spouse_pension_amount',
                'spouse_pension_id','spouse_pension_issuer'},
    },
    'spouse_other_income': {
        'gid':guuid(6),'cat':'spouse','sort':1,'rep':True,
        'label':'Sonstige Einkünfte des Ehepartners','key':'spouse_other_income',
        'keys':{'spouse_other_income_type','spouse_other_income_amount'},
    },
}
KEY_TO_GRP = {}
for gname, gdata in GROUPS.items():
    for k in gdata['keys']:
        KEY_TO_GRP[k] = gname

# ── Visibility-rule overrides (co-founder verified) ──────────────────────────
# Keys whose dependency was missing or wrong in the source Excel, or where a
# section-level condition is needed.  Keyed by question key; value replaces
# whatever parse_dep() returned.

MARITAL_STATUSES_WITH_PARTNER = [
    "eheähnliche Gemeinschaft",
    "eingetragene Lebenspartnerschaft",
    "verheiratet",
    "dauernd getrennt lebend",
]

VIS_OVERRIDES: dict[str, Optional[dict]] = {
    # Spouse section — show only when a legal/civil partner exists.
    # Applied to top-level spouse questions (those with no existing dep).
    # Sub-questions that already reference a spouse-specific answer keep their dep.
    # Overrides are applied AFTER parse_dep; only questions where parse_dep returns
    # None AND key starts with "spouse_" get this applied automatically (see below).

    # Wealth — missing deps from Excel:
    'bank_savings_account_amount': {'question_key': 'bank_savings_account_yes_no', 'value': 'Ja'},
    'bank_savings_iban':           {'question_key': 'bank_savings_account_yes_no', 'value': 'Ja'},
    'life_insurance_total_amount': {'question_key': 'life_insurance',              'not_value': 'Nein'},
    'life_insurance_name':         {'question_key': 'life_insurance',              'not_value': 'Nein'},
    'life_insurance_number':       {'question_key': 'life_insurance',              'not_value': 'Nein'},

    # Expenditure — spouse civil-servant sub-questions missing dep:
    'spouse_health_insurance_amount': {'question_key': 'spouse_govermental_employee', 'value': 'Ja'},
    'spouse_care_insurance_amount':   {'question_key': 'spouse_govermental_employee', 'value': 'Ja'},
}

SPOUSE_TOP_LEVEL_VIS = {
    'question_key': 'marital_status',
    'in_values': MARITAL_STATUSES_WITH_PARTNER,
}

# ── Read questions ────────────────────────────────────────────────────────────
wb_q = openpyxl.load_workbook(QUESTIONS_FILE)
ws_q = wb_q['Berlin_HzP']

questions = []
for row in ws_q.iter_rows(min_row=2, values_only=True):
    xid,section,key,prompt = row[0],row[1],row[2],row[3]
    xtype,opts,dep,comment  = row[5],row[6],row[7],row[8]
    if not xid or not section:
        continue
    xid = int(xid)
    section = SECTIONS.get(section)
    if section is None:
        continue          # General → auto, skip
    atype = TYPE_MAP.get(xtype)
    if atype is None:
        continue          # auto-populated field, skip
    # Normalise key
    key = KEY_FIX.get(key, key)
    # Rename duplicate keys
    if xid == 97  and key == 'automobile_owner':       key = 'automobile_holder'
    if xid == 158 and key == 'spouse_automobile_owner': key = 'spouse_automobile_holder'
    # Determine final visibility rule
    vis = VIS_OVERRIDES.get(key) or parse_dep(dep)
    # Top-level spouse questions with no dep → apply partner-status condition
    if section == 'spouse' and vis is None:
        vis = SPOUSE_TOP_LEVEL_VIS

    questions.append({
        'id': xid, 'section': section, 'key': key,
        'prompt': prompt or '', 'atype': atype,
        'opts': opts, 'vis': vis, 'comment': comment or '',
    })

# ── Read care homes ───────────────────────────────────────────────────────────
wb_c = openpyxl.load_workbook(CARE_HOMES_FILE)
ws_c = wb_c['care_homes_seed']
care_homes = []
for row in ws_c.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    name   = row[2] or ''
    street = (str(row[5] or '') + ' ' + str(row[6] or '')).strip()
    plz    = str(int(row[7])).zfill(5) if row[7] else ''
    city   = row[8] or ''
    addr   = f'{street}, {plz} {city}'.strip(', ')
    care_homes.append({'name': name, 'address': addr})

# ── Read Berlin PLZs ──────────────────────────────────────────────────────────
wb_p = openpyxl.load_workbook(PLZ_FILE)
ws_p = wb_p['deutschland_plz_sozialamt_named']
berlin_plzs = sorted({
    str(int(row[0])).zfill(5)
    for row in ws_p.iter_rows(min_row=2, values_only=True)
    if row[0] and row[1] and 'Berlin' in str(row[1])
})

# ── Generate SQL ──────────────────────────────────────────────────────────────
lines = []
L = lines.append

L("-- ============================================================")
L("-- supabase/seed.sql")
L("-- Hilfe-zur-Pflege — reproducible config seed, Milestone 2")
L("--")
L("-- Source files:")
L("--   seed-care homes.xlsx  (7 partner care homes)")
L("--   Question Master_Berlin.xlsx → Berlin_HzP sheet")
L("--   plz_de.xlsx → deutschland_plz_sozialamt_named sheet (Berlin only)")
L("--")
L("-- Run automatically by:  supabase db reset  /  supabase start")
L("-- ============================================================")
L("")
L("-- ─── Clear old seed data (safe order: children first) ──────")
L("DELETE FROM public.document_rule;")
L("DELETE FROM public.document_type;")
L("DELETE FROM public.question_option;")
L("DELETE FROM public.question;")
L("DELETE FROM public.question_group;")
L("DELETE FROM public.category;")
L("DELETE FROM public.questionnaire;")
L("DELETE FROM public.postal_code_rule;")
L("DELETE FROM public.care_home;")
L("DELETE FROM public.social_office;")
L("")

# ── Social office ─────────────────────────────────────────────────────────────
L("-- ─── Social office ──────────────────────────────────────────")
L("-- One canonical entry for Berlin; PLZ routing resolves to this office.")
L("-- Per-borough detail (Bezirksamt) can be added when per-borough forms are ready.")
L("INSERT INTO public.social_office (id, name, address, contact_email, contact_phone) VALUES")
L(f"  ('{suuid(1)}', 'Sozialamt Berlin',")
L( "   'Bezirksämter Berlin – Amt für Soziales',")
L( "   'buergerbuero@sozialamt.berlin.de', '030 115')")
L("ON CONFLICT DO NOTHING;")
L("")

# ── Care homes ────────────────────────────────────────────────────────────────
L("-- ─── Care homes (7 partner homes) ──────────────────────────")
L("INSERT INTO public.care_home (id, name, address) VALUES")
ch = [f"  ('{cuuid(i+1)}', {sq(h['name'])}, {sq(h['address'])})"
      for i, h in enumerate(care_homes)]
L(',\n'.join(ch))
L("ON CONFLICT DO NOTHING;")
L("")

# ── Questionnaires ────────────────────────────────────────────────────────────
L("-- ─── Questionnaires ─────────────────────────────────────────")
L("INSERT INTO public.questionnaire (id, social_office_id, name, version) VALUES")
L(f"  ('{quuid(1)}', '{suuid(1)}', 'Fragebogen – Sozialamt Berlin', 1),")
L(f"  ('{quuid(2)}', NULL,         'Allgemeiner Fragebogen (Fallback)', 1)")
L("ON CONFLICT DO NOTHING;")
L("")

# ── Categories ────────────────────────────────────────────────────────────────
L("-- ─── Categories ─────────────────────────────────────────────")
L("INSERT INTO public.category (id, questionnaire_id, key, sort_order, label_de) VALUES")
cats = [
    f"  ('{CAT_ID[s]}', '{quuid(1)}', {sq(s)}, {CAT_SORT[s]}, {sq(CAT_LABEL[s])})"
    for s in sorted(CAT_SORT, key=CAT_SORT.get)
]
L(',\n'.join(cats))
L("ON CONFLICT DO NOTHING;")
L("")

# ── Question groups ───────────────────────────────────────────────────────────
L("-- ─── Question groups (repeatable) ──────────────────────────")
L("INSERT INTO public.question_group")
L("  (id, category_id, key, sort_order, label_de, is_repeatable, min_count, max_count)")
L("VALUES")
grps = [
    f"  ('{g['gid']}', '{CAT_ID[g['cat']]}', {sq(g['key'])}, {g['sort']}, {sq(g['label'])}, true, 0, NULL)"
    for g in GROUPS.values()
]
L(',\n'.join(grps))
L("ON CONFLICT DO NOTHING;")
L("")

# ── Questions ─────────────────────────────────────────────────────────────────
L("-- ─── Questions ──────────────────────────────────────────────")
L("-- Prompts are all about the PATIENT (care-home resident), not the caregiver.")
L("-- No question IDs or per-office logic appear in UI components.")
L("INSERT INTO public.question")
L("  (id, category_id, group_id, key, sort_order, answer_type,")
L("   is_required, prompt_de, help_de, validation, visibility_rule)")
L("VALUES")
cat_counters = {s: 0 for s in CAT_SORT}
qrows = []
for r in questions:
    sec  = r['section']
    so   = cat_counters[sec]; cat_counters[sec] += 1
    cat_id = CAT_ID[sec]
    grp_name = KEY_TO_GRP.get(r['key'])
    grp_id_sql = f"'{GROUPS[grp_name]['gid']}'" if grp_name else 'NULL'
    vis = r['vis']
    qrows.append(
        f"  ('{qquid(r['id'])}', '{cat_id}', {grp_id_sql},\n"
        f"   {sq(r['key'])}, {so}, '{r['atype']}', true,\n"
        f"   {sq(r['prompt'])}, NULL, NULL, {jb(vis)})"
    )
L(',\n'.join(qrows))
L("ON CONFLICT DO NOTHING;")
L("")

# ── Options ───────────────────────────────────────────────────────────────────
L("-- ─── Question options ───────────────────────────────────────")
L("INSERT INTO public.question_option (id, question_id, key, sort_order, label_de, value) VALUES")
opt_rows = []
for r in questions:
    if not r['opts'] or r['atype'] not in ('single_select', 'multi_select'):
        continue
    opts = [o.strip() for o in str(r['opts']).split(',') if o.strip()]
    for i, opt in enumerate(opts):
        # Make a safe SQL key: lower, replace non-alphanumeric with _
        ok = re.sub(r'[^a-z0-9_]', '_', opt.lower())
        ok = re.sub(r'_+', '_', ok).strip('_')[:50]
        opt_rows.append(
            f"  ('{ouuid(r['id'], i)}', '{qquid(r['id'])}', {sq(ok)}, {i}, {sq(opt)}, {sq(opt)})"
        )
L(',\n'.join(opt_rows))
L("ON CONFLICT DO NOTHING;")
L("")

# ── PLZ rules ─────────────────────────────────────────────────────────────────
L(f"-- ─── PLZ rules — {len(berlin_plzs)} Berlin codes ──────────────────────────")
L("-- Each PLZ is an individual rule (plz_from = plz_to).")
L("-- Unrecognised PLZs will fall to the fallback questionnaire in M3.")
L("INSERT INTO public.postal_code_rule (id, social_office_id, plz_from, plz_to, priority) VALUES")
plz_rows = [
    f"  ('{plzuuid(i+1)}', '{suuid(1)}', '{plz}', '{plz}', 10)"
    for i, plz in enumerate(berlin_plzs)
]
L(',\n'.join(plz_rows))
L("ON CONFLICT DO NOTHING;")
L("")

print('\n'.join(lines))

# ── Stats to stderr ───────────────────────────────────────────────────────────
print(f"\n-- STATS --", file=sys.stderr)
print(f"Questions   : {len(questions)}", file=sys.stderr)
print(f"Care homes  : {len(care_homes)}", file=sys.stderr)
print(f"PLZ rules   : {len(berlin_plzs)}", file=sys.stderr)
cats_q = {}
for r in questions:
    cats_q.setdefault(r['section'],[]).append(r['key'])
for s in sorted(CAT_SORT, key=CAT_SORT.get):
    qs = cats_q.get(s,[])
    print(f"  {s:<15} {len(qs):3d} questions", file=sys.stderr)
